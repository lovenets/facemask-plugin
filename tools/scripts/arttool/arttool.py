# ==============================================================================
# Copyright (C) 2017 General Workings Inc
#
# This program is free software; you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation; either version 2 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program; if not, write to the Free Software
# Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston, MA  02110-1301, USA
# ==============================================================================


# ==============================================================================
# IMPORTS
# ==============================================================================
import sys, subprocess, os, json, uuid, time
from shutil import copyfile
from copy import deepcopy
import tempfile
from PyQt5.QtWidgets import QApplication, QWidget, QLabel, QListWidget, QVBoxLayout, QTabWidget
from PyQt5.QtWidgets import QPushButton, QComboBox, QDateTimeEdit, QDialogButtonBox, QMessageBox
from PyQt5.QtWidgets import QScrollArea, QMainWindow, QCheckBox, QHBoxLayout, QTextEdit, QFileDialog
from PyQt5.QtWidgets import QLineEdit, QFrame, QDialog, QFrame, QSplitter
from PyQt5.QtGui import QIcon, QBrush, QColor, QFont, QPixmap, QMovie
from PyQt5.QtCore import QDateTime, Qt

"""
from PIL import Image as PILImage

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font
"""

from .utils import *
from .releases import *
from .additions import *

# don't check svn more often than this
SVN_CHECK_TIME = (60 * 5)  # 5 minutes is lots

# ==============================================================================
# MAIN WINDOW : ArtToolWindow class
# ==============================================================================

FIELD_WIDTH = 180
PANE_WIDTH = 690
TEXTURE_SIZES = ["32", "64", "128", "256", "512", "1024", "2048", "4096", "8192", "16384"]
CATEGORIES = ["Top", "Eyes", "Ears", "Nose", "Mouth", "Neck", "Full", "Combo", "Other"]
TIERS = ["1", "2", "3"]
MASK_UI_NAMES = {"name": "Pretty Name",
                 "description": "Description",
                 "author": "Author",
                 "tags": "Tags",
                 "category": "Category",
                 "tier": "Tier (1,2 or 3)",
                 "depth_head": "Depth Head",
                 "is_morph": "Morph Mask",
                 "is_vip": "V.I.P. Mask",
                 "is_intro": "Intro Animation",
                 "release_with_plugin": "Release With Plugin",
                 "do_not_release": "DO NOT RELEASE",
                 "texture_max": "Max Texture Size",
                 "intro_fade_time": "Intro Fade Time",
                 "intro_duration": "Intro Duration",
                 "license": "License",
                 "website": "Website"}

MASK_UI_FIELDS = ["name", "description", "author", "tags", "category", "tier", "depth_head",
                  "is_morph", "is_vip", "is_intro", "release_with_plugin", "do_not_release",
                  "texture_max", "intro_fade_time", "intro_duration", "license", "website"]
COMBO_UI_FIELDS = ["name", "description", "author", "tags", "tier", "is_vip", "is_intro",
                   "release_with_plugin", "do_not_release", "texture_max", "intro_fade_time",
                   "intro_duration", "license", "website"]
UI_FIELDS = [MASK_UI_FIELDS, COMBO_UI_FIELDS]


MASK_FIELD_TOOLTIPS = {"tier": "The tier of the mask.\nTier 1 is most expensive, 3 is least expensive.",
                       "tags": "Comma separated list of tags.",
                       "category": "The area of the face the mask covers.",
                       "depth_head": "Whether this mask needs a depth occlusion head added.",
                       "is_morph": "This FBX is a morph FBX",
                       "is_vip": "VIP mask for a specific streamer.",
                       "is_intro": "This mask is used as an intro or outro animation.",
                       "do_not_release": "Check this box if the public should never see this."}

DROP_DOWNS = {"texture_max": TEXTURE_SIZES,
              "category": CATEGORIES,
              "tier" : TIERS}

RELEASE_FIELDS = ["name", "uuid", "description", "author", "tags", "category", "tier", "is_vip", "is_intro"]


class ArtToolWindow(QMainWindow):

    # --------------------------------------------------
    # CONSTRUCTOR
    # --------------------------------------------------
    def __init__(self, *args):
        super(ArtToolWindow, self).__init__(*args)

        # State
        self.metadata = None
        self.comboTabIdx = -1
        self.currentFbx = -1
        self.currentCombo = -1
        self.ignoreSVN = 0
        self.additionsClipboard = None
        self.cancelledSVN = False
        self.editPane = None
        self.currentFilter = None
        self.lastSVNCheck = 0
        self.dialogUp = False
        self.mainLayout = None

        # Load our config
        self.config = createGetConfig()

        # Left Pane
        leftPane = QWidget()
        leftLayout = QVBoxLayout(leftPane)

        # Streamlabs logo
        slabslogo = QLabel()
        slabslogo.setPixmap(QPixmap("arttool/streamlabs.png"))
        slabslogo.setScaledContents(True)
        slabslogo.show()
        leftLayout.addWidget(slabslogo)
        slabslogo.setMaximumWidth(300)
        slabslogo.setMaximumHeight(53)

        # Filter box
        self.fbxfilter = QLineEdit()
        self.fbxfilter.editingFinished.connect(lambda: self.onFbxFilterChanged())
        leftLayout.addWidget(self.fbxfilter)

        # make a list widget for fbx's
        self.fbxlist = QListWidget()
        self.fbxlist.itemSelectionChanged.connect(lambda: self.onFbxClicked())
        self.fbxlist.setMinimumHeight(640)

        # make the combo tab
        combotab = QWidget()
        combotab.setAutoFillBackground(True)

        # combo buttons
        b = QPushButton("Add")
        b.setParent(combotab)
        b.setGeometry(5,5, 75, 30)
        b.pressed.connect(lambda: self.onAddCombo())
        b = QPushButton("Del")
        b.setParent(combotab)
        b.setGeometry(85, 5, 75, 30)
        b.pressed.connect(lambda: self.onDelCombo())

        # make alist widget for combos
        self.combolist = QListWidget()
        self.combolist.itemSelectionChanged.connect(lambda: self.onComboClicked())
        self.combolist.setParent(combotab)
        self.combolist.setGeometry(0, 40, 294, 600)

        # make fbx/combo tabs
        tabs = QTabWidget()
        tabs.currentChanged.connect(lambda i: self.onComboTabChanged(i))
        tabs.setGeometry(0,0,300, 640)
        tabs.addTab(self.fbxlist, "FBX")
        tabs.addTab(combotab, "COMBOS")

        leftLayout.addWidget(tabs)

        # top splitter
        topSplitter = QSplitter(Qt.Horizontal)
        topSplitter.addWidget(leftPane)

        # Layout for edit pane
        rightPane = QWidget()
        self.mainLayout = QHBoxLayout(rightPane)
        topSplitter.addWidget(rightPane)

        # Fill in files lists
        self.fillFbxList()
        self.fillComboList()

        # crate main splitter
        mainSplitter = QSplitter(Qt.Vertical)
        mainSplitter.addWidget(topSplitter)

        # bottom pane
        bottomPane = QWidget()
        bottomArea = QHBoxLayout(bottomPane)

        # output window
        self.outputWindow = QTextEdit()
        self.outputWindow.setMinimumHeight(90)
        bottomArea.addWidget(self.outputWindow)

        # buttons area
        buttonArea = QWidget()
        buttonArea.setMinimumWidth(150)
        buttonArea.setMinimumHeight(60)
        locs = [(0, 0), (0, 30), (0, 60), (75, 0), (75, 30), (75, 60)]
        c = 0
        for nn in ["Refresh", "Autobuild", "Rebuild All", "S3 Upload", "XLS Export", "Release Masks"]:
            b = QPushButton(nn)
            b.setParent(buttonArea)
            (x, y) = locs[c]
            c += 1
            b.setGeometry(x, y, 75, 30)
            b.pressed.connect(lambda nn=nn: self.onMainButton(nn))
        bottomArea.addWidget(buttonArea)

        mainSplitter.addWidget(bottomPane)

        # Show the window
        self.setCentralWidget(mainSplitter)
        self.setGeometry(self.config["x"], self.config["y"], 1024, 900)
        self.setWindowTitle('Streamlabs Art Tool')
        self.setWindowIcon(QIcon('arttool/arttoolicon.png'))

        # Check our binaries
        self.checkBinaries()

    # --------------------------------------------------
    # Fill(refill) in the fbx list
    # --------------------------------------------------
    def fillFbxList(self):
        # Get list of fbx files
        self.fbxfiles = getFbxFileList(".")

        # Clear list
        while self.fbxlist.count() > 0:
            i = self.fbxlist.takeItem(0)
            i = None

        # Get filter
        filt = self.fbxfilter.text().lower()
        if not filt or len(filt) < 1:
            filt = None

        # Fill list
        self.fbxlistMap = dict()
        self.fbxlistRevMap = dict()
        fbxidx = 0
        listidx = 0
        for fbx in self.fbxfiles:
            if not filt or filt in fbx.lower():
                self.fbxlist.addItem(fbx[2:])
                self.fbxlistMap[listidx] = fbxidx
                self.fbxlistRevMap[fbxidx] = listidx
                listidx += 1
            fbxidx += 1

        # color fbxlist items
        for idx in range(0, len(self.fbxfiles)):
            self.setFbxColorIcon(idx)

        self.resetEditPane()


    # --------------------------------------------------
    # Fill(refill) in the combos list
    # --------------------------------------------------
    def fillComboList(self):
        # Get list of fbx files
        self.combofiles = getComboFileList(".")

        # Clear list
        while self.combolist.count() > 0:
            i = self.combolist.takeItem(0)
            i = None

        # Fill list
        for combo in self.combofiles:
            self.combolist.addItem(combo[2:])
            idx = self.combolist.count() - 1

        # color list items
        for idx in range(0, len(self.combofiles)):
            self.setComboColorIcon(idx)

        self.resetEditPane()


    def resetEditPane(self):
        self.currentCombo = -1
        self.currentFbx = -1
        self.metadata = None
        self.fbxlist.clearSelection()
        self.combolist.clearSelection()
        if self.mainLayout is not None:
            self.createMaskEditPane(None)

    # --------------------------------------------------
    # Check binaries
    # --------------------------------------------------
    def checkBinaries(self):
        gotSVN = os.path.exists(SVNBIN.replace('"', ''))
        gotMM = os.path.exists(MASKMAKERBIN)
        gotRP = os.path.exists(MORPHRESTFILE)

        if not gotSVN:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setText("You seem to be missing " + os.path.basename(SVNBIN))
            msg.setInformativeText("You should (re)install tortoiseSVN, and be sure to install the command line tools.")
            msg.setWindowTitle("Missing Binary File")
            msg.setStandardButtons(QMessageBox.Ok)
            self.ignoreSVN += 1
            self.dialogUp = True
            msg.exec_()
            self.dialogUp = False
        if not gotMM:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setText("You seem to be missing " + os.path.basename(MASKMAKERBIN))
            msg.setWindowTitle("Missing Binary File")
            msg.setStandardButtons(QMessageBox.Ok)
            self.ignoreSVN += 1
            self.dialogUp = True
            msg.exec_()
            self.dialogUp = False
        if not gotRP:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setText("You seem to be missing " + os.path.basename(MORPHRESTFILE))
            msg.setWindowTitle("Missing Binary File")
            msg.setStandardButtons(QMessageBox.Ok)
            self.ignoreSVN += 1
            self.dialogUp = True
            msg.exec_()
            self.dialogUp = False

    # --------------------------------------------------
    # Create generic widget
    # --------------------------------------------------
    def createLabelWidget(self, parent, name, field, x, y, noedit=False):
        q = QLabel(name)
        q.setParent(parent)
        q.setGeometry(x, y, FIELD_WIDTH - 10, 26)
        q.setFont(QFont("Arial", 12, QFont.Bold))
        q.show()

        if field == None:
            return q

        critical = critical_mask
        if self.metadata["fbx"].lower().endswith(".json"):
            critical = critical_combo

        value = self.metadata[field]

        if field in DROP_DOWNS:
            dropvals = DROP_DOWNS[field]
            q = QComboBox()
            idx = 0
            selidx = -1
            for s in dropvals:
                if s == str(value):
                    selidx = idx
                q.addItem(s)
                idx += 1
            if selidx < 0:
                selidx = 0

            q.setCurrentIndex(selidx)
            q.currentIndexChanged.connect(lambda state: self.onDropdownChanged(state, field))
        elif type(value) is str or type(value) is float or type(value) is int:
            if noedit:
                q = QLabel(str(value))
            else:
                q = QLineEdit(str(value))
                q.textChanged.connect(lambda text: self.onTextFieldChanged(text, field))
            if critical(field) and len(value) == 0:
                q.setStyleSheet("border: 1px solid #FF0000;")
            elif desired(field) and len(value) == 0:
                q.setStyleSheet("border: 1px solid #FF7F50;")
            else:
                q.setStyleSheet("border: 0px;")
        elif type(value) is bool:
            q = QCheckBox()
            q.setChecked(value)
            q.stateChanged.connect(lambda state: self.onCheckboxChanged(state, field))

        q.setParent(parent)
        if type(value) is int:
            q.setGeometry(x + FIELD_WIDTH - 10, y, FIELD_WIDTH, 20)
        else:
            q.setGeometry(x + FIELD_WIDTH - 10, y, PANE_WIDTH - FIELD_WIDTH - 10, 20)
        q.setFont(QFont("Arial", 12, QFont.Bold))
        q.show()
        return q

    # --------------------------------------------------
    # Colors and Icons for main FBX list
    # --------------------------------------------------
    def setFbxColorIconInternal(self, mdc, mt, nb, idx):
        self.fbxlist.item(idx).setFont(QFont("Arial", 12, QFont.Bold))
        if mdc == CHECKMETA_GOOD:
            self.fbxlist.item(idx).setForeground(QBrush(QColor("#32CD32")))
        elif mdc == CHECKMETA_ERROR:
            self.fbxlist.item(idx).setForeground(QBrush(QColor("#FF0000")))
        elif mdc == CHECKMETA_WARNING:
            self.fbxlist.item(idx).setForeground(QBrush(QColor("#FF7F50")))
        elif mdc == CHECKMETA_NORELEASE:
            self.fbxlist.item(idx).setForeground(QBrush(QColor("#000000")))
        elif mdc == CHECKMETA_WITHPLUGIN:
            self.fbxlist.item(idx).setForeground(QBrush(QColor("#5070FF")))
        if mt == MASK_UNKNOWN:
            self.fbxlist.item(idx).setIcon(QIcon("arttool/unknownicon.png"))
        else:
            if nb:
                if mt == MASK_NORMAL:
                    self.fbxlist.item(idx).setIcon(QIcon("arttool/maskicon_build.png"))
                elif mt == MASK_MORPH:
                    self.fbxlist.item(idx).setIcon(QIcon("arttool/morphicon_build.png"))
            else:
                if mt == MASK_NORMAL:
                    self.fbxlist.item(idx).setIcon(QIcon("arttool/maskicon.png"))
                elif mt == MASK_MORPH:
                    self.fbxlist.item(idx).setIcon(QIcon("arttool/morphicon.png"))

    def setFbxColorIcon(self, idx):
        if idx in self.fbxlistRevMap:
            mdc, mt = checkMetaDataFile(self.fbxfiles[idx])
            nb = doesFileNeedRebuilding(self.fbxfiles[idx])
            self.setFbxColorIconInternal(mdc, mt, nb, self.fbxlistRevMap[idx])

    def setComboColorIconInternal(self, mdc, mt, nb, idx):
        self.combolist.item(idx).setFont(QFont("Arial", 12, QFont.Bold))
        if mdc == CHECKMETA_GOOD:
            self.combolist.item(idx).setForeground(QBrush(QColor("#32CD32")))
        elif mdc == CHECKMETA_ERROR:
            self.combolist.item(idx).setForeground(QBrush(QColor("#FF0000")))
        elif mdc == CHECKMETA_WARNING:
            self.combolist.item(idx).setForeground(QBrush(QColor("#FF7F50")))
        elif mdc == CHECKMETA_NORELEASE:
            self.combolist.item(idx).setForeground(QBrush(QColor("#000000")))
        elif mdc == CHECKMETA_WITHPLUGIN:
            self.combolist.item(idx).setForeground(QBrush(QColor("#5070FF")))
        if nb:
            self.combolist.item(idx).setIcon(QIcon("arttool/comboicon_build.png"))
        else:
            self.combolist.item(idx).setIcon(QIcon("arttool/comboicon.png"))

    def setComboColorIcon(self, idx):
        mdc, mt = checkMetaDataFile(self.combofiles[idx])
        nb = doesFileNeedRebuilding(self.combofiles[idx])
        self.setComboColorIconInternal(mdc, mt, nb, idx)


    def updateListColorIcon(self):
        mdc, mt = checkMetaData(self.metadata)
        nb = doesFileNeedRebuilding(self.metadata["fbx"], self.metadata)
        if self.comboTabIdx == 0:
            self.setFbxColorIconInternal(mdc, mt, nb, self.fbxlistRevMap[self.currentFbx])
        else:
            self.setComboColorIconInternal(mdc, mt, nb, self.currentCombo)



    # --------------------------------------------------
    # createMaskEditPane
    # --------------------------------------------------
    def createMaskEditPane(self, fbxfile):
        if self.editPane:
            self.mainLayout.removeWidget(self.editPane)
            self.editPane.deleteLater()

        self.editPane = QWidget()
        self.mainLayout.addWidget(self.editPane)
        self.editPane.setMinimumWidth(PANE_WIDTH)
        self.editPane.show()

        # empty pane
        if fbxfile == None:
            return

        # mask icon png
        q = QLabel()
        q.setParent(self.editPane)
        pf = os.path.abspath(fbxfile.lower().replace(".fbx", ".gif").replace(".json", ".gif"))
        if os.path.exists(pf):
            m = QMovie(pf)
        else:
            m = QMovie("arttool/noicon.png")
        q.setMovie(m)
        m.start()
        q.setScaledContents(True)
        q.setGeometry(0, 2, 256, 256)
        q.show()

        # mask file name
        q = QLabel(fbxfile[2:])
        q.setParent(self.editPane)
        q.setGeometry(260, 44, 600, 36)
        q.setFont(QFont("Arial", 14, QFont.Bold))
        q.show()

        # uuid
        q = QLabel(self.metadata["uuid"])
        q.setParent(self.editPane)
        q.setGeometry(260, 80, 600, 20)
        q.setStyleSheet("font: 10pt;")
        # q.setFont(QFont( "Arial", 6))
        q.show()

        # buttons
        b = QPushButton("BUILD")
        b.setParent(self.editPane)
        b.setGeometry(260, 2, 64, 32)
        q.setFont(QFont("Arial", 14, QFont.Bold))
        b.pressed.connect(lambda: self.onBuild())
        b.show()

        # Tabbed Panel
        tabs = QTabWidget(self.editPane)
        tabs.setGeometry(0, 264, PANE_WIDTH, 600)

        # mask meta data fields
        tab1 = QWidget()
        y = 2
        dy = 28
        self.paneWidgets = dict()
        for field in UI_FIELDS[self.comboTabIdx]:
            if field != "uuid":
                w = self.createLabelWidget(tab1, MASK_UI_NAMES[field], field, 10, y)
                self.paneWidgets[field] = w
                if field in MASK_FIELD_TOOLTIPS:
                    w.setToolTip(MASK_FIELD_TOOLTIPS[field])
                y += dy
        tab1.setAutoFillBackground(True)
        tabs.addTab(tab1, "Mask Meta Data")

        if self.comboTabIdx == 0:
            tab2 = self.createAdditionsTab()
            tabs.addTab(tab2, "Additions")
        else:
            tab2 = self.createCombosTab()
            tabs.addTab(tab2, "Combinations")

        tabs.show()


    def createCombosTab(self):
        # combos tab
        tab2 = QWidget()
        y = 10
        dy = 40
        self.createLabelWidget(tab2, "Combo Files", None, 10, y)
        y += dy

        self.comboWidgets = list()

        for cidx in range(0,10):
            q = QComboBox()
            idx = 1
            selidx = -1
            value = self.metadata["additions"][cidx]
            q.addItem(" None ")
            for f in self.fbxfiles:
                if f == str(value):
                    selidx = idx
                q.addItem(f[2:])
                idx += 1
            if selidx < 0:
                selidx = 0
            q.setCurrentIndex(selidx)
            q.setParent(tab2)
            q.setGeometry(10, y, 600, 30)
            y += dy
            q.setFont(QFont("Arial", 12, QFont.Bold))
            q.show()
            q.currentIndexChanged.connect(lambda state, cidx=cidx: self.onComboFileChanged(state,cidx))
            self.comboWidgets.append(q)

        tab2.setAutoFillBackground(True)
        return tab2


    def createAdditionsTab(self):
        # additions tab
        tab2 = QWidget()
        y = 10
        dy = 40
        self.createLabelWidget(tab2, "Mask Build Additions", None, 10, y)
        y += dy

        # make a list widget
        self.addslist = QListWidget()
        if "additions" in self.metadata:
            additions = self.metadata["additions"]
            idx = 0
            for addition in additions:
                self.addslist.addItem(addition["type"] + " : " + addition["name"])
                self.addslist.item(idx).setFont(QFont("Arial", 12, QFont.Bold))
                idx += 1
        self.addslist.itemDoubleClicked.connect(lambda: self.onEditAddition())
        self.addslist.setParent(tab2)
        self.addslist.setGeometry(10, y, PANE_WIDTH - 20, 250)
        y += 260

        # bottom buttons for additons
        x = 10
        b = QPushButton("Add")
        b.setParent(tab2)
        b.setGeometry(x, y, 75, 30)
        b.pressed.connect(lambda: self.onAddAddition())
        x += 85
        b = QPushButton("Edit")
        b.setParent(tab2)
        b.setGeometry(x, y, 75, 30)
        b.pressed.connect(lambda: self.onEditAddition())
        x += 85
        b = QPushButton("Del")
        b.setParent(tab2)
        b.setGeometry(x, y, 75, 30)
        b.pressed.connect(lambda: self.onDelAddition())
        x += 85
        b = QPushButton("Dupe")
        b.setParent(tab2)
        b.setGeometry(x, y, 75, 30)
        b.pressed.connect(lambda: self.onCopyAddition())
        x += 85
        b = QPushButton("Up")
        b.setParent(tab2)
        b.setGeometry(x, y, 75, 30)
        b.pressed.connect(lambda: self.onMoveUpAddition())
        x += 85
        b = QPushButton("Down")
        b.setParent(tab2)
        b.setGeometry(x, y, 75, 30)
        b.pressed.connect(lambda: self.onMoveDownAddition())

        # top buttons for additions
        x = 300
        b = QPushButton("COPY ALL")
        b.setParent(tab2)
        b.setGeometry(x, 10, 75, 30)
        b.pressed.connect(lambda: self.onCopyAllAdditions())
        x += 85
        b = QPushButton("PASTE ALL")
        b.setParent(tab2)
        b.setGeometry(x, 10, 75, 30)
        b.pressed.connect(lambda: self.onPasteAllAdditions())

        tab2.setAutoFillBackground(True)
        return tab2


    # --------------------------------------------------
    # saveCurrentMetadata
    # --------------------------------------------------
    def saveCurrentMetadata(self):
        if self.metadata:
            fbxfile = None
            if self.currentFbx >= 0:
                fbxfile = self.fbxfiles[self.currentFbx]
            elif self.currentCombo >= 0:
                fbxfile = self.combofiles[self.currentCombo]

            if fbxfile:
                metafile = getMetaFileName(fbxfile)
                oldmetadata = loadMetadataFile(fbxfile)
                if oldmetadata != self.metadata:
                    writeMetaData(metafile, self.metadata, True)
                    print("saving", metafile)

    # --------------------------------------------------
    # WIDGET SIGNALS CALLBACKS
    # --------------------------------------------------

    # Main buttons
    def onMainButton(self, button):
        if button == "S3 Upload":
            self.uploadToS3()
        elif button == "Refresh":
            self.fillFbxList()
            self.fillComboList()
        elif button == "Autobuild":
            self.doAutobuild()
        elif button == "Rebuild All":
            self.doRebuildAll()
        elif button == "XLS Export":
            self.onWriteMetadataExcel()
        elif button == "Release Masks":
            self.doReleaseMasks()

    # TODO : hook up to a button
    def onMakeThumbsMovie(self):
        # self.ignoreSVN += 1
        # self.dialogUp = True
        # addn = ReleasesDialog.go_modal(self)
        # self.dialogUp = False

        # quick hack for fun
        framenum = 1
        for fbx in self.fbxfiles:
            src = os.path.abspath(fbx.replace(".fbx", ".png").replace(".FBX", ".png"))
            # dst = "c:\\Temp\\masks\\frame%04d.png" % framenum
            dst = "c:\\Temp\\masks\\" + os.path.basename(src)

            fileOkay = True
            if "frogHead" in src:
                fileOkay = False
            if "joshog" in src:
                fileOkay = False

            if os.path.exists(src) and fileOkay:
                print("copy", src, dst)
                framenum += 1
                copyfile(src, dst)


    def onComboTabChanged(self, tab):
        self.comboTabIdx = tab
        self.resetEditPane()


    def onComboFileChanged(self, state, which):
        if state == 0:
            self.metadata["additions"][which] = ""
        else:
            self.metadata["additions"][which] = self.fbxfiles[state - 1]


    # FBX file clicked in list
    def onFbxClicked(self):
        if self.comboTabIdx != 0:
            return
        k = self.fbxlist.currentRow()
        if k >= 0 and k < self.fbxlist.count():
            self.saveCurrentMetadata()
            self.currentCombo = -1
            self.currentFbx = self.fbxlistMap[k]
            fbxfile = self.fbxfiles[self.currentFbx]
            self.metadata = createGetMetaData(fbxfile)
            self.updateListColorIcon()
            self.createMaskEditPane(fbxfile)


    # FBX Filter box changed
    def onFbxFilterChanged(self):
        filt = self.fbxfilter.text().lower()
        if not filt or len(filt) < 1:
            filt = None
        if filt != self.currentFilter:
            self.fillFbxList()
            self.currentFilter = filt


    # FBX file clicked in list
    def onComboClicked(self):
        if self.comboTabIdx != 1:
            return
        k = self.combolist.currentRow()
        if k >= 0 and k < self.combolist.count():
            self.saveCurrentMetadata()
            self.currentCombo = k
            self.currentFbx = -1
            combofile = self.combofiles[self.currentCombo]
            self.metadata = createGetMetaData(combofile)
            self.updateListColorIcon()
            self.createMaskEditPane(combofile)


    def onAddCombo(self):
        file, filter = QFileDialog.getSaveFileName(self, 'Save file', os.path.abspath("."),
                                                   "Mask files (*.json)")
        if file is not None and len(file) > 0:
            combofile = make_path_relative(os.path.abspath("."), os.path.abspath(file))
            print("adding combo", combofile)
            self.saveCurrentMetadata()
            metadata = createGetMetaData(combofile)
            self.fillComboList()
            for idx in range(0, len(self.combofiles)):
                if combofile == self.combofiles[idx]:
                    self.currentCombo = idx
                    break
            if self.currentCombo >= 0:
                self.combolist.setCurrentRow(self.currentCombo)


    def onDelCombo(self):
        pass


    # text field changed
    def onTextFieldChanged(self, text, field):
        if type(self.metadata[field]) is int:
            self.metadata[field] = int(text)
        elif type(self.metadata[field]) is float:
            self.metadata[field] = float(text)
        else:
            self.metadata[field] = text

        critical = critical_mask
        if self.metadata["fbx"].lower().endswith(".json"):
            critical = critical_combo

        if critical(field) and len(text) == 0:
            self.paneWidgets[field].setStyleSheet("border: 1px solid #FF0000;")
        elif desired(field) and len(text) == 0:
            self.paneWidgets[field].setStyleSheet("border: 1px solid #FF7F50;")
        else:
            self.paneWidgets[field].setStyleSheet("border: 0px;")
        self.updateListColorIcon()


    # checkbox changed
    def onCheckboxChanged(self, state, field):
        if state == 0:
            self.metadata[field] = False
        else:
            self.metadata[field] = True
        self.updateListColorIcon()


    # texsize changed
    def onDropdownChanged(self, state, field):
        if type(self.metadata[field]) is int:
            self.metadata[field] = int(DROP_DOWNS[field][state])
        elif type(self.metadata[field]) is float:
            self.metadata[field] = float(DROP_DOWNS[field][state])
        else:
            self.metadata[field] = DROP_DOWNS[field][state]


    # called before exit
    def finalCleanup(self):
        self.cancelledSVN = True
        self.saveCurrentMetadata()

        # This is just getting annoying
        #
        # QApplication.setOverrideCursor(Qt.WaitCursor)
        #needcommit = svnNeedsCommit()
        # QApplication.restoreOverrideCursor()
        #if needcommit:
        #    msg = QMessageBox()
        #    msg.setIcon(QMessageBox.Warning)
        #    msg.setText("You have changed files in your depot.")
        #    msg.setInformativeText("Be sure to commit your changes to avoid conflicts.")
        #    msg.setWindowTitle("Files Changed - SVN Commit Recommended")
        #    msg.setStandardButtons(QMessageBox.Ok)
        #    self.ignoreSVN += 1
        #    self.dialogUp = True
        #    msg.exec_()
        #    self.dialogUp = False

        #
        # WRITE CONFIG
        #
        geo = self.geometry()
        self.config["x"] = geo.x()
        self.config["y"] = geo.y()
        writeMetaData(getConfigFile(), self.config)


    # build
    def onBuild(self):

        QApplication.setOverrideCursor(Qt.WaitCursor)

        if self.currentCombo >= 0:
            # build combo
            combofile = self.combofiles[self.currentCombo]
            buildCombo(combofile, self.outputWindow, self.metadata)
            filetype = "Combo Json"

        else:
            # build mask
            fbxfile = self.fbxfiles[self.currentFbx]
            svnAddFile(fbxfile)
            buildMask(fbxfile, self.outputWindow, self.metadata)
            filetype = "FBX"

        deps, missing = getDependencies(self.metadata)

        for d in deps:
            svnAddFile(d["file"])

        for m in missing:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setText("This " + filetype +" depends on " + m + ", which cannot be found.")
            msg.setWindowTitle("Missing PNG File")
            msg.setStandardButtons(QMessageBox.Ok)
            self.ignoreSVN += 1
            self.dialogUp = True
            msg.exec_()
            self.dialogUp = False

        QApplication.restoreOverrideCursor()

        self.updateListColorIcon()


    def onAddAddition(self):
        self.ignoreSVN += 1
        self.dialogUp = True
        addn = NewAdditionDialog.go_modal(self)
        self.dialogUp = False
        if addn:
            if "additions" not in self.metadata:
                self.metadata["additions"] = list()
            self.metadata["additions"].append(addn)
            idx = self.addslist.count()
            self.addslist.addItem(addn["type"] + " : " + addn["name"])
            self.addslist.item(idx).setFont(QFont("Arial", 12, QFont.Bold))


    def onEditAddition(self):
        idx = self.addslist.currentRow()
        if idx >= 0:
            self.ignoreSVN += 1
            self.dialogUp = True
            addn = AdditionDialog.go_modal(self, self.metadata["additions"][idx])
            self.dialogUp = False
            if addn:
                self.addslist.item(idx).setText(addn["type"] + " : " + addn["name"])
                self.metadata["additions"][idx] = addn


    def onDelAddition(self):
        idx = self.addslist.currentRow()
        if idx >= 0:
            del self.metadata["additions"][idx]
            i = self.addslist.takeItem(idx)
            i = None


    def onCopyAddition(self):
        idx = self.addslist.currentRow()
        if idx >= 0:
            addn = deepcopy(self.metadata["additions"][idx])
            self.addslist.insertItem(idx + 1, addn["type"] + " : " + addn["name"])
            self.addslist.item(idx + 1).setFont(QFont("Arial", 12, QFont.Bold))
            self.addslist.setCurrentRow(idx + 1)
            self.metadata["additions"].insert(idx + 1, addn)


    def onMoveUpAddition(self):
        idx = self.addslist.currentRow()
        if idx > 0:
            self.addslist.insertItem(idx - 1, self.addslist.takeItem(idx))
            self.addslist.setCurrentRow(idx - 1)
            self.metadata["additions"].insert(idx - 1, self.metadata["additions"].pop(idx))


    def onMoveDownAddition(self):
        idx = self.addslist.currentRow()
        if idx >= 0 and self.addslist.count() > 1 and idx < (self.addslist.count() - 1):
            self.addslist.insertItem(idx + 1, self.addslist.takeItem(idx))
            self.addslist.setCurrentRow(idx + 1)
            self.metadata["additions"].insert(idx + 1, self.metadata["additions"].pop(idx))


    def onCopyAllAdditions(self):
        self.additionsClipboard = deepcopy(self.metadata["additions"])


    def onPasteAllAdditions(self):
        if self.additionsClipboard:
            self.metadata["additions"].extend(deepcopy(self.additionsClipboard))
            for addn in self.additionsClipboard:
                idx = self.addslist.count()
                self.addslist.addItem(addn["type"] + " : " + addn["name"])
                self.addslist.item(idx).setFont(QFont("Arial", 12, QFont.Bold))


    def onFocusChanged(self, old, now):

        if self.cancelledSVN or self.dialogUp:
            return

        if not old and now:

            if self.ignoreSVN > 0:
                self.ignoreSVN -= 1
                return

            if (time.time() - self.lastSVNCheck) < SVN_CHECK_TIME:
                return

            QApplication.setOverrideCursor(Qt.WaitCursor)
            needsup = svnNeedsUpdate()
            QApplication.restoreOverrideCursor()
            self.lastSVNCheck = time.time()
            if needsup:
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Warning)
                msg.setText("Your SVN repository is out of date.")
                msg.setInformativeText("Would you like to sync up now?")
                msg.setWindowTitle("SVN Repository out of date")
                msg.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
                msg.buttonClicked.connect(lambda i: self.onSyncOk(i))
                self.ignoreSVN += 1
                self.dialogUp = True
                msg.exec_()
                self.dialogUp = False


    def doSVNUpdate(self):
        QApplication.setOverrideCursor(Qt.WaitCursor)
        arttoolUpdated = svnUpdate(self.outputWindow)
        self.fillFbxList()
        QApplication.restoreOverrideCursor()

        if arttoolUpdated:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setText("File(s) in the Art Tool have changed.")
            msg.setInformativeText("You should exit the Art Tool now and start it again.")
            msg.setWindowTitle("Art Tool Changed")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.buttonClicked.connect(lambda i: self.onSyncOk(i))
            self.cancelledSVN = True
            msg.exec_()

    def onSyncOk(self, i):
        if i.text() == "OK":
            self.doSVNUpdate()
        else:
            # dont check anymore
            self.cancelledSVN = True

    def doAutobuild(self):

        QApplication.setOverrideCursor(Qt.WaitCursor)

        all_missing = dict()

        for f in self.fbxfiles:
            if doesFileNeedRebuilding(f):
                deps, missing = buildMask(f, self.outputWindow)
                if len(missing) > 0:
                    all_missing[f] = missing

        for f in self.combofiles:
            if doesFileNeedRebuilding(f):
                deps, missing = buildCombo(f, self.outputWindow)
                if len(missing) > 0:
                    all_missing[f] = missing

        for file, missing in all_missing.items():
            for m in missing:
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Warning)
                msg.setText(file +" depends on " + m + ", which cannot be found.")
                msg.setWindowTitle("Missing PNG File")
                msg.setStandardButtons(QMessageBox.Ok)
                self.ignoreSVN += 1
                self.dialogUp = True
                msg.exec_()
                self.dialogUp = False

        QApplication.restoreOverrideCursor()
        self.fillComboList()
        self.fillFbxList()


    def doRebuildAll(self):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Warning)
        msg.setText("WARNING!! This will rebuild ALL THE MASKS!!")
        msg.setInformativeText("Are you abso-fucking-lutely sure?")
        msg.setWindowTitle("Rebuild Everything")
        msg.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
        msg.buttonClicked.connect(lambda i: self.onRebuildAllOk(i))
        self.ignoreSVN += 1
        self.dialogUp = True
        msg.exec_()
        self.dialogUp = False


    def onRebuildAllOk(self, i):
        if i.text() == "OK":

            QApplication.setOverrideCursor(Qt.WaitCursor)

            all_missing = dict()

            for f in self.fbxfiles:
                deps, missing = buildMask(f, self.outputWindow)
                if len(missing) > 0:
                    all_missing[f] = missing

            for f in self.combofiles:
                deps, missing = buildCombo(f, self.outputWindow)
                if len(missing) > 0:
                    all_missing[f] = missing

            for file, missing in all_missing.items():
                for m in missing:
                    msg = QMessageBox()
                    msg.setIcon(QMessageBox.Warning)
                    msg.setText(file + " depends on " + m + ", which cannot be found.")
                    msg.setWindowTitle("Missing PNG File")
                    msg.setStandardButtons(QMessageBox.Ok)
                    self.ignoreSVN += 1
                    self.dialogUp = True
                    msg.exec_()
                    self.dialogUp = False

            QApplication.restoreOverrideCursor()
            self.fillComboList()
            self.fillFbxList()

        else:
            # dont check anymore
            self.cancelledSVN = True

    def doReleaseMasks(self):
        print("todo release masks")

    def uploadToS3(self):

        # what now
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Warning)
        msg.setText("Upload to S3, or just build index?")
        msg.setWindowTitle("Upload to S3")
        msg.setStandardButtons(QMessageBox.Ok|QMessageBox.Cancel)
        msg.buttonClicked.connect(lambda i: self.onUploadToS3Confirm(i))
        self.ignoreSVN += 1
        self.dialogUp = True
        msg.exec_()
        self.dialogUp = False


    def onUploadToS3Confirm(self, item):

        do_upload = item.text() == "OK"

        metalist = list()
        jsonlist = list()
        for fbxfile in self.fbxfiles:
            mdc, mt = checkMetaDataFile(fbxfile)
            if mdc == CHECKMETA_GOOD:
                metadata = loadMetadataFile(fbxfile)
                d = dict()
                for k in RELEASE_FIELDS:
                    d[k] = metadata[k]
                    if type(d[k]) is str:
                        d[k] = d[k].replace("\n", "").replace("\r", "")
                metalist.append(d)
                jsonlist.append(jsonFromFbx(fbxfile))

        for combofile in self.combofiles:
            mdc, mt = checkMetaDataFile(combofile)
            if mdc == CHECKMETA_GOOD:
                metadata = loadMetadataFile(combofile)
                d = dict()
                for k in RELEASE_FIELDS:
                    d[k] = metadata[k]
                    if type(d[k]) is str:
                        d[k] = d[k].replace("\n", "").replace("\r", "")
                md = getCombinedComboMeta(metadata)
                d["tags"] = md["tags"]
                d["author"] = md["author"]

                metalist.append(d)
                jsonlist.append(combofile)

        for i in range(0,len(metalist)):
            metalist[i]["category"] = metalist[i]["category"].lower()
            metalist[i]["tags"] = metalist[i]["tags"].lower().replace(", ",",")
            metalist[i]["modtime"] = int(os.path.getmtime(jsonlist[i]))
            metalist[i]["author"] = metalist[i]["author"].replace(", ",",")

            if metalist[i]["tags"].endswith(" "):
                metalist[i]["tags"] = metalist[i]["tags"][:-1]
            metalist[i]["name"] = metalist[i]["name"].strip()
            metalist[i]["author"] = metalist[i]["author"].strip()

            if do_upload:
                print("Uploading",jsonlist[i])

                uuid = metalist[i]["uuid"]
                print("  json...")
                s3_upload(jsonlist[i], uuid + ".json")
                print("  png...")
                s3_upload(jsonlist[i].replace(".json",".png"), uuid + ".png")
                print("  gif...")
                s3_upload(jsonlist[i].replace(".json",".gif"), uuid + ".gif")
                print("  mp4...")
                s3_upload(jsonlist[i].replace(".json",".mp4"), uuid + ".mp4")


        file, filter = QFileDialog.getSaveFileName(self, 'Save file', os.path.abspath("."),
                                                   "Mask files (*.json)")
        if file is not None and len(file) > 0:
            writeMetaData(os.path.abspath(file), metalist)


    def onWriteMetadataExcel(self):

        pass

        """

        QApplication.setOverrideCursor(Qt.WaitCursor)

        file, filter = QFileDialog.getSaveFileName(self, 'Save file', os.path.abspath("."),
                                                   "Excel (*.xlsx)")
        if file is not None and len(file) > 0:
            outfile = os.path.abspath(file)
        else:
            return

        wb = Workbook()
        ws = wb.active

        HEADER_VALS = { "A1": "Icon",
                        "B1": "Flags",
                        "C1": "Name",
                        "D1": "Description",
                        "E1": "Tags",
                        "F1": "Category",
                        "G1": "Tier",
                        "H1": "UUID",
                        "I1": "Author"}

        for c,v in HEADER_VALS.items():
            ws[c] = v
            ws[c].font = Font(bold=True, size=16)
        ws.row_dimensions[1].height = 50

        row = 2
        allfiles = self.fbxfiles
        allfiles.extend(self.combofiles)
        for fbxfile in self.fbxfiles:

            pngfile = fixpath(fbxfile.lower().replace(".fbx",".png").replace(".json",".png"))
            metadata = loadMetadataFile(fbxfile)
            if metadata:
                mdc, mt = checkMetaData(metadata)
            else:
                mdc = CHECKMETA_ERROR

            if mdc == CHECKMETA_GOOD:
                pimg = PILImage.open(pngfile).resize((64, 64), PILImage.LANCZOS)
                tf = tempfile.NamedTemporaryFile(suffix=".png")
                pimg.save(tf)
                img = Image(tf)
                ws.add_image(img, "A"+str(row))
                ws.row_dimensions[row].height = 50

                ff = ""
                if metadata["is_morph"]:
                    ff += " MORPH "
                if metadata["is_intro"]:
                    ff += " INTRO "
                if metadata["is_vip"]:
                    ff += " VIP "
                if metadata["category"] == "Combo":
                    ff += " COMBO "
                ws['B'+str(row)] = ff

                ws['C'+str(row)] = metadata["name"]
                ws['D'+str(row)] = metadata["description"]
                ws['E'+str(row)] = metadata["tags"]
                ws['F'+str(row)] = metadata["category"]
                ws['G'+str(row)] = metadata["tier"]
                ws['H'+str(row)] = metadata["uuid"]
                ws['I'+str(row)] = metadata["author"]

                row += 1

        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 45
        ws.column_dimensions["E"].width = 45
        ws.column_dimensions["F"].width = 15
        ws.column_dimensions["G"].width = 8
        ws.column_dimensions["H"].width = 38
        ws.column_dimensions["I"].width = 35

        wb.save(outfile)
        tf.close()
        wb.close()

        QApplication.restoreOverrideCursor()
        """